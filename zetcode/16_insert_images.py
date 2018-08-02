#!/usr/bin/env python3

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

book = Workbook()
sheet = book.active

sheet.merge_cells('A1:F2')
cell = sheet.cell(row=1, column=1)
cell.value = 'Python 3'
cell.alignment = Alignment(horizontal='center', vertical='center')


img = Image('Python-Logo-PNG-Image.png')
sheet.add_image(img, 'A3')

book.save('sheet_image.xlsx')
print('Done.')