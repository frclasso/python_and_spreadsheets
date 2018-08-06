#!/usr/bin/env python3

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(df=True, index=True, header=True):
    ws.append(r)

for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'


wb.save('pandas_openpy.xlsx')
print('Done.')