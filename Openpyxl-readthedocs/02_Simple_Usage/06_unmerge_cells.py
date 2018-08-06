#!/usr/bin/env python3

from openpyxl import load_workbook
wb = load_workbook('merged.xlsx')
ws = wb.active

ws.unmerge_cells('A2:D2')
#or
ws.unmerge_cells(start_row=4, start_column=2, end_row=8, end_column=8)

wb.save('unmerged.xlsx')
print('Done.')