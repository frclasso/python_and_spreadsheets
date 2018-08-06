#!/usr/bin/env python3

from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws.merge_cells('A2:D2')

#or

ws.merge_cells(start_row=4, start_column=2, end_row=8, end_column=8)

wb.save('merged.xlsx')
print('Done.')