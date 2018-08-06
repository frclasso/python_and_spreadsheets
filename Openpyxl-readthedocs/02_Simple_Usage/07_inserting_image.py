#!/usr/bin/env python3

from openpyxl import Workbook
from openpyxl.drawing.image import Image

wb = Workbook()
ws = wb.active

ws.merge_cells('A1:F2')
ws['A1'] = 'You should to see a Pyhon logo below'

img = Image('py_logo.png')
ws.add_image(img, 'A3')
wb.save('image.xlsx')
print('Done.')