#!/usr/bin/env python3

from  openpyxl import Workbook
from  openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = 'range names'

for row in range(1, 20):
    ws1.append(range(601))

ws2 = wb.create_sheet(title="Pi")
ws2["F5"] = 3.14159

ws3 = wb.create_sheet(title='Data')
for row in range(10, 20):
    for col in range(7, 34):
        _= ws3.cell(column=col, row=row, value='{0}'.format(get_column_letter(col)))
print(ws3["G10"].value)

wb.save(filename=dest_filename)