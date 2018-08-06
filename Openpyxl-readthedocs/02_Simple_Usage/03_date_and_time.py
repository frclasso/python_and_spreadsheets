#!/usr/bin/env python3

import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws['A1'] = datetime.datetime(2010, 7, 21)
print(ws['A1'].number_format)

wb.guess_types=True

ws["B1"] = '3.14%'
wb.guess_types=False
print(ws['B1'].value)
print(ws['B1'].number_format)