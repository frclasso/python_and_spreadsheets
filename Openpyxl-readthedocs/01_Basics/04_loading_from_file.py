#!/usr/bin/env python3

from openpyxl import load_workbook
wb = load_workbook('teste.xlsx')
print(wb.sheetnames)