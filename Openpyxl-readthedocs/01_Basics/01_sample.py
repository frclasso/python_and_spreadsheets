#!/usr/bin/env python3

from openpyxl import Workbook
wb = Workbook()
ws = wb.active # grab the active worksheet

# Data can assigned directly to cells
ws['A1'] = 42

# Rowns can also appended
ws.append([1,2,3])

# Python types will automatically converted
import datetime
ws['B1'] = datetime.datetime.now()

wb.save('sample.xlsx')
print('Done.')