#!/usr/bin/env python3

from openpyxl import Workbook
wb = Workbook()

ws =wb.active

ws1 = wb.create_sheet('DadosPaciente', 0)
ws2 = wb.create_sheet('Charts',1)
ws3 = wb.create_sheet('DadosCientificos',2)

ws1.sheet_properties.tabColor = '1072BA'  # alterando cor de fundo
ws3.sheet_properties.tabColor = '1072BA'  # alterando cor de fundo

# ws.title ="New title"  # para renomear a Sheet original

wb.remove_sheet(wb["Sheet"])  # Poderia alterar o nome ao inves de deletar
print(wb.sheetnames)


# for sheet in wb:
#     print(sheet.title)

# Criando copias da worksheet

# source = wb.active
# #source = ws2
# target = wb.copy_worksheet(source)

ws1['A1'] = 'indices'
ws1['B1'] = 'Dados do paciente'
ws1['C1'] = 'referencia'

ws2['A1'] = 'indices'
ws2['B1'] = 'Dados do paciente'
ws2['C1'] = 'referencia'

ws3['A1'] = 'indices'
ws3['B1'] = 'Dados do paciente'
ws3['C1'] = 'referencia'

# Podemos definir valores usando notacao row/column diretamente
#ws1.cell(row=4,column=2, value=10)
ws2.cell(row=4,column=2, value='Charts')


print(tuple(ws1.rows))

wb.save('teste.xlsx')
print('Done')