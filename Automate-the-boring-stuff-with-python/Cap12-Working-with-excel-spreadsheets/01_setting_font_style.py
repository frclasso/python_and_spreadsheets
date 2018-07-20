#!/usr/bin/env python3

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, Border, Side

wb = Workbook()
ws = wb.active
sheet = wb.get_sheet_by_name('Sheet')
highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=12, italic=True)
bd = Side(style="thick", color="000000")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
wb.add_named_style(highlight)

ws['A1'].style = 'highlight'
ws['B1'].style = 'highlight'
ws['C1'].style = 'highlight'

sheet["A1"] = "Produto"
sheet["B1"] = "Valor"
sheet["C1"] = "Quantidade"

wb.save('exemplo3.xlsx')
print('Done.')
