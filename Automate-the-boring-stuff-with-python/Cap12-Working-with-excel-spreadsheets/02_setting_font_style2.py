#!/usr/bin/env python3

import openpyxl

import os
os.chdir("/home/fabio/Desktop/estudo_ti/Python/python_and_spreadsheets/Automate-the-boring-stuff-with-python")

from openpyxl.workbook import Workbook
from openpyxl.styles import Font, NamedStyle

wb = Workbook()
ws = wb.active
sheet = wb.get_sheet_by_name('Sheet')

fontObj1 = NamedStyle("fontObj1")
fontObj1.font = Font(name="Times New Roman", bold=True)
wb.add_named_style(fontObj1)

ws['A1'].style = "fontObj1"
sheet["A1"] = "Bold Times New Roman"

fontObj2 = NamedStyle("fontObj2")
fontObj2.font = Font(size=24, italic=True, name="Arial", color="000111")
wb.add_named_style(fontObj2)

ws['B3'].style = 'fontObj2'
sheet["B3"] = '24pt Italic'

wb.save("styles.xlsx")
print("Done.")