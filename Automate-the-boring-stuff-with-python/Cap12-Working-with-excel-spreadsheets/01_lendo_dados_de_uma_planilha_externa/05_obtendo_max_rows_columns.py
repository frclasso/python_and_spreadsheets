#!/usr/bin/env python3

import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
print("Quantidade de linhas:", sheet.max_row)
print("Quantidade de colunas:",sheet.max_column)