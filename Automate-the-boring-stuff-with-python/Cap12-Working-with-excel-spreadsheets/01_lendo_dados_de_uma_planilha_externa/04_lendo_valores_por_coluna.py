#!/usr/bin/env python3

"""Nesse exemplo obteremos os valores da coluna B (2) utilizando um loop do tipo for()"""

import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
#print(sheet.cell(row=1, column=2))

# Imprimindo o primeiro valor
#print(sheet.cell(row=1, column=2).value)

# Imprimindo todos os valores, utilizando um loop for()
for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)

