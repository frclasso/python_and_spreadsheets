#!/usr/bin/env python3

"""Conversao entre as letras das colunas para numeros
   Para converter de letras para numeros utilizamos a funcao : openpyxl.cell.column_index_from_string()
   Para converter numeros para letras utilizamos a funcao: openpyxl.cell.get.column_letter()
"""

import openpyxl

from openpyxl.utils import get_column_letter,  column_index_from_string

# Fornecendo o indice numerico para obter o indice alfabetico
print(get_column_letter(1)) # A
print(get_column_letter(2)) # B
print(get_column_letter(27)) # AA
print(get_column_letter(900)) # AHP

print()
wb = openpyxl.load_workbook("example.xlsx")
sheet = wb.get_sheet_by_name("Sheet1")
# Obtendo o numero maximo de colunas
print(get_column_letter(sheet.max_column)) # C
#Obtendo o indice das colunas
print(column_index_from_string('A'))
print(column_index_from_string('AA'))
