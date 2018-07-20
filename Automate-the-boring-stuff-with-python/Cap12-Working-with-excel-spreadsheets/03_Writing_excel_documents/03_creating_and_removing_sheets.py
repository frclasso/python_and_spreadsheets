#!/usr/bin/env python3

"""Sheets can be added to and removed from a workbook with the create_sheet()
and remove_sheet() methods"""

import openpyxl

wb = openpyxl.Workbook()
#print(wb.get_sheet_names())
wb.create_sheet()
#print(wb.get_sheet_names())  # ['Sheet', 'Sheet1']
wb.create_sheet(index=0, title='First Sheet')
#print(wb.get_sheet_names())  # ['First Sheet','Sheet', 'Sheet1']
wb.create_sheet(index=2, title='Middle Sheet')
#print(wb.get_sheet_names())  # ['First Sheet','Sheet','Middle Sheet', 'Sheet1']

# Removendo sheets
wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
wb.remove_sheet(wb.get_sheet_by_name('Sheet1'))
print(wb.get_sheet_names())
