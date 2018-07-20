#!/usr/bin/env python3

""" if you want to see the result of the calculation for the formula
instead of the literal formula, you must pass True for the data_only keyword
argument to load_workbook(). This means a Workbook object can show either
the formulas or the result of the formulas but not both. """

import openpyxl

wbFormulas = openpyxl.load_workbook('writeFormula.xlsx')
sheet = wbFormulas.get_active_sheet()
print(sheet["A3"].value)

wbDataOnly = openpyxl.load_workbook('writeFormula.xlsx', data_only=True)
sheet = wbDataOnly.get_active_sheet()
print(sheet["A3"].value) # deu None, era pra dar 500, a soma.

