#!/usr/bin/env python3

"""A rectangular area of cells can be merged into a single cell with the
merge_cells() sheet method"""

import openpyxl

wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells.'

wb.save('merged.xlsx')
print('Done.')